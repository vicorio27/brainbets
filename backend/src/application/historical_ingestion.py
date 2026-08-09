"""Historical data ingestion service.

Downloads public datasets and normalizes them into the PostgreSQL schema.
Currently supports:
- Tennis: tennis-data.co.uk ATP match archives (XLS inside ZIP)
- Football: football-data.co.uk league season CSVs
"""
import csv
import io
import zipfile
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from uuid import uuid4

import requests
import urllib3
import xlrd
from openpyxl import load_workbook
from sqlalchemy.orm import Session

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


def _read_excel(path: Path) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Read both .xls (xlrd) and .xlsx (openpyxl) files."""
    if path.suffix.lower() == ".xls":
        workbook = xlrd.open_workbook(str(path))
        sheet = workbook.sheet_by_index(0)
        headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
        rows = []
        for row_idx in range(1, sheet.nrows):
            rows.append({headers[col]: sheet.cell_value(row_idx, col) for col in range(sheet.ncols)})
        return headers, rows
    else:
        workbook = load_workbook(str(path), data_only=True)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: row[i] for i in range(len(headers))})
        return headers, rows

from src.domain.models import (
    Competitor,
    IngestionJob,
    League,
    Match,
    MatchCompetitor,
    MatchScore,
    Sport,
)


CACHE_DIR = Path("/tmp/historical_raw")

# football-data.co.uk league code -> name mapping (major leagues)
FOOTBALL_LEAGUE_NAMES = {
    "E0": "Premier League",
    "E1": "Championship",
    "E2": "League One",
    "E3": "League Two",
    "SP1": "La Liga",
    "SP2": "Segunda Division",
    "I1": "Serie A",
    "I2": "Serie B",
    "D1": "Bundesliga",
    "D2": "2. Bundesliga",
    "F1": "Ligue 1",
    "F2": "Ligue 2",
    "N1": "Eredivisie",
    "B1": "Jupiler Pro League",
    "P1": "Primeira Liga",
    "T1": "Süper Lig",
    "G1": "Super League Greece",
}


def _excel_date_to_datetime(value: Any) -> Optional[datetime]:
    """Convert Excel serial date, datetime, or string to datetime."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value
    if isinstance(value, (int, float)):
        # Excel serial date (Windows 1900 epoch)
        base = datetime(1899, 12, 30, tzinfo=timezone.utc)
        return base + timedelta(days=float(value))
    if isinstance(value, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value, fmt).replace(tzinfo=timezone.utc)
            except ValueError:
                continue
    return None


def _download(url: str, cache_path: Path, force: bool = False, max_retries: int = 3) -> bytes:
    cache_path.parent.mkdir(parents=True, exist_ok=True)

    def _fetch() -> bytes:
        try:
            resp = requests.get(url, timeout=120)
            resp.raise_for_status()
        except requests.exceptions.SSLError:
            resp = requests.get(url, timeout=120, verify=False)
            resp.raise_for_status()
        return resp.content

    if not force and cache_path.exists():
        data = cache_path.read_bytes()
        if cache_path.suffix.lower() == ".zip":
            try:
                with zipfile.ZipFile(io.BytesIO(data)) as z:
                    z.namelist()
                return data
            except zipfile.BadZipFile:
                pass
        else:
            return data

    last_error = None
    for attempt in range(max_retries):
        try:
            data = _fetch()
            if cache_path.suffix.lower() == ".zip":
                with zipfile.ZipFile(io.BytesIO(data)) as z:
                    z.namelist()
            cache_path.write_bytes(data)
            return data
        except Exception as e:
            last_error = e
            if attempt < max_retries - 1:
                import time
                time.sleep(2 ** attempt)
    raise last_error or RuntimeError(f"Failed to download {url}")


def _ensure_sport(db: Session) -> Sport:
    sport = db.query(Sport).filter(Sport.code == "football").first()
    if not sport:
        sport = Sport(id=uuid4(), code="football", name="Football")
        db.add(sport)
        db.commit()
        db.refresh(sport)
    return sport


def _ensure_tennis_sport(db: Session) -> Sport:
    sport = db.query(Sport).filter(Sport.code == "tennis").first()
    if not sport:
        sport = Sport(id=uuid4(), code="tennis", name="Tennis")
        db.add(sport)
        db.commit()
        db.refresh(sport)
    return sport


def _ensure_league(db: Session, sport_id: Any, name: str, external_id: Optional[str] = None) -> League:
    filters = [League.sport_id == sport_id, League.name == name]
    if external_id:
        filters.append(League.external_id == external_id)
    league = db.query(League).filter(*filters).first()
    if not league:
        league = League(id=uuid4(), sport_id=sport_id, name=name, external_id=external_id)
        db.add(league)
        db.commit()
        db.refresh(league)
    return league


def _ensure_competitor(db: Session, sport_id: Any, name: str, competitor_type: str) -> Competitor:
    competitor = (
        db.query(Competitor)
        .filter(
            Competitor.sport_id == sport_id,
            Competitor.name.ilike(name),
            Competitor.type == competitor_type,
        )
        .first()
    )
    if not competitor:
        competitor = Competitor(
            id=uuid4(),
            sport_id=sport_id,
            name=name,
            type=competitor_type,
        )
        db.add(competitor)
        db.commit()
        db.refresh(competitor)
    return competitor


def _match_external_id_tennis(row: Dict[str, Any]) -> str:
    """Build a stable external_id for a tennis match."""
    date_val = row.get("Date", "")
    return f"TENNIS-{date_val}-{row.get('Winner', '')}-{row.get('Loser', '')}".replace(" ", "_")


def _match_external_id_football(season: str, league: str, row: Dict[str, Any]) -> str:
    """Build a stable external_id for a football match."""
    date = row.get("Date", "").replace("/", "")
    home = row.get("HomeTeam", "").replace(" ", "_")
    away = row.get("AwayTeam", "").replace(" ", "_")
    return f"FOOTBALL-{season}-{league}-{date}-{home}-{away}"


def ingest_tennis_season(
    db: Session,
    year: int,
    force_download: bool = False,
) -> Dict[str, Any]:
    """Ingest one ATP season from tennis-data.co.uk.

    2010-2020 are distributed as .zip files containing an Excel file.
    2021+ are distributed as direct .xlsx files.
    """
    xls_path = None

    # Try .zip first (legacy years)
    zip_url = f"http://www.tennis-data.co.uk/{year}/{year}.zip"
    zip_path = CACHE_DIR / f"tennis_{year}.zip"
    try:
        zip_bytes = _download(zip_url, zip_path, force=force_download)
        with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
            xls_members = [m for m in z.namelist() if m.lower().endswith((".xls", ".xlsx"))]
            if not xls_members:
                raise ValueError(f"No Excel file found in {zip_path}")
            xls_member = xls_members[0]
            z.extract(xls_member, CACHE_DIR)
            xls_path = CACHE_DIR / xls_member
    except (zipfile.BadZipFile, ValueError, requests.exceptions.HTTPError, RuntimeError):
        # Fall back to direct .xlsx download
        if zip_path.exists():
            zip_path.unlink()
        xlsx_url = f"http://www.tennis-data.co.uk/{year}/{year}.xlsx"
        xlsx_path = CACHE_DIR / f"tennis_{year}.xlsx"
        _download(xlsx_url, xlsx_path, force=force_download)
        xls_path = xlsx_path

    headers, rows = _read_excel(xls_path)

    sport = _ensure_tennis_sport(db)
    inserted = 0
    updated = 0
    errors = 0
    season = str(year)

    for row in rows:
        try:
            tournament = str(row.get("Tournament", "")).strip()
            if not tournament:
                continue

            league = _ensure_league(db, sport.id, tournament)
            match_date = _excel_date_to_datetime(row.get("Date"))
            if not match_date:
                continue

            winner_name = str(row.get("Winner", "")).strip()
            loser_name = str(row.get("Loser", "")).strip()
            if not winner_name or not loser_name:
                continue

            winner = _ensure_competitor(db, sport.id, winner_name, "player")
            loser = _ensure_competitor(db, sport.id, loser_name, "player")

            external_id = _match_external_id_tennis(row)
            match = (
                db.query(Match)
                .filter(Match.external_id == external_id, Match.sport_id == sport.id)
                .first()
            )

            surface = str(row.get("Surface", "")).strip() or None
            round_name = str(row.get("Round", "")).strip() or None
            best_of = int(row.get("Best of", 3)) if row.get("Best of") else 3

            if not match:
                match = Match(
                    id=uuid4(),
                    sport_id=sport.id,
                    league_id=league.id,
                    external_id=external_id,
                    source_api="tennis-data.co.uk",
                    match_date=match_date,
                    status="FINISHED",
                    season=season,
                    extra_data={
                        "surface": surface,
                        "round": round_name,
                        "best_of": best_of,
                        "series": str(row.get("Series", "")).strip(),
                        "court": str(row.get("Court", "")).strip(),
                    },
                )
                db.add(match)
                db.commit()
                db.refresh(match)
                inserted += 1
            else:
                match.match_date = match_date
                match.status = "FINISHED"
                match.season = season
                match.extra_data = {
                    **(match.extra_data or {}),
                    "surface": surface,
                    "round": round_name,
                    "best_of": best_of,
                }
                db.commit()
                updated += 1

            # Upsert competitors links
            for side, competitor, rank_key, pts_key in [
                ("player1", winner, "WRank", "WPts"),
                ("player2", loser, "LRank", "LPts"),
            ]:
                link = (
                    db.query(MatchCompetitor)
                    .filter(MatchCompetitor.match_id == match.id, MatchCompetitor.side == side)
                    .first()
                )
                rank = row.get(rank_key)
                pts = row.get(pts_key)
                extra = {}
                if pts:
                    extra["points"] = pts
                if not link:
                    link = MatchCompetitor(
                        id=uuid4(),
                        match_id=match.id,
                        competitor_id=competitor.id,
                        side=side,
                        pre_match_ranking=int(rank) if rank and str(rank).isdigit() else None,
                        extra_data=extra,
                    )
                    db.add(link)
                else:
                    link.pre_match_ranking = int(rank) if rank and str(rank).isdigit() else link.pre_match_ranking
                    link.extra_data = {**(link.extra_data or {}), **extra}
                db.commit()

            # Score: tennis-data.co.uk provides Wsets/Lsets columns (winner/loser sets)
            try:
                winner_sets_raw = row.get("Wsets", "")
                loser_sets_raw = row.get("Lsets", "")
                if winner_sets_raw != "" and loser_sets_raw != "":
                    winner_sets = int(float(winner_sets_raw))
                    loser_sets = int(float(loser_sets_raw))
                    score = (
                        db.query(MatchScore)
                        .filter(MatchScore.match_id == match.id, MatchScore.period == "FULL_TIME")
                        .first()
                    )
                    if not score:
                        score = MatchScore(
                            id=uuid4(),
                            match_id=match.id,
                            home_score=winner_sets,
                            away_score=loser_sets,
                            period="FULL_TIME",
                        )
                        db.add(score)
                    else:
                        score.home_score = winner_sets
                        score.away_score = loser_sets
                    db.commit()
            except Exception:
                pass
        except Exception as e:
            db.rollback()
            errors += 1

    return {"inserted": inserted, "updated": updated, "errors": errors}


def ingest_football_season(
    db: Session,
    season: str,
    league: str,
    force_download: bool = False,
) -> Dict[str, Any]:
    """Ingest one football season/league from football-data.co.uk."""
    url = f"https://www.football-data.co.uk/mmz4281/{season}/{league}.csv"
    csv_path = CACHE_DIR / f"football_{season}_{league}.csv"

    text = _download(url, csv_path, force=force_download).decode("utf-8", errors="replace")
    lines = [line for line in text.splitlines() if line.strip()]
    reader = csv.DictReader(lines)
    rows = list(reader)

    sport = _ensure_sport(db)
    league_name = FOOTBALL_LEAGUE_NAMES.get(league, f"League {league}")
    league_obj = _ensure_league(db, sport.id, league_name, external_id=league)

    inserted = 0
    updated = 0
    errors = 0

    for row in rows:
        try:
            home_team = str(row.get("HomeTeam", "")).strip()
            away_team = str(row.get("AwayTeam", "")).strip()
            if not home_team or not away_team:
                continue

            match_date = _excel_date_to_datetime(row.get("Date"))
            if not match_date:
                continue

            external_id = _match_external_id_football(season, league, row)
            match = (
                db.query(Match)
                .filter(Match.external_id == external_id, Match.sport_id == sport.id)
                .first()
            )

            if not match:
                match = Match(
                    id=uuid4(),
                    sport_id=sport.id,
                    league_id=league_obj.id,
                    external_id=external_id,
                    source_api="football-data.co.uk",
                    match_date=match_date,
                    status="FINISHED",
                    season=season,
                    extra_data={
                        "season": season,
                        "league_code": league,
                        "referee": row.get("Referee"),
                    },
                )
                db.add(match)
                db.commit()
                db.refresh(match)
                inserted += 1
            else:
                match.match_date = match_date
                match.status = "FINISHED"
                match.season = season
                match.extra_data = {
                    **(match.extra_data or {}),
                    "season": season,
                    "league_code": league,
                    "referee": row.get("Referee"),
                }
                db.commit()
                updated += 1

            home = _ensure_competitor(db, sport.id, home_team, "team")
            away = _ensure_competitor(db, sport.id, away_team, "team")

            for side, competitor, rank_key in [
                ("home", home, None),
                ("away", away, None),
            ]:
                link = (
                    db.query(MatchCompetitor)
                    .filter(MatchCompetitor.match_id == match.id, MatchCompetitor.side == side)
                    .first()
                )
                if not link:
                    link = MatchCompetitor(
                        id=uuid4(),
                        match_id=match.id,
                        competitor_id=competitor.id,
                        side=side,
                    )
                    db.add(link)
                db.commit()

            # Full-time score
            try:
                home_score = int(float(row.get("FTHG", 0))) if row.get("FTHG") else None
                away_score = int(float(row.get("FTAG", 0))) if row.get("FTAG") else None
                if home_score is not None and away_score is not None:
                    score = (
                        db.query(MatchScore)
                        .filter(MatchScore.match_id == match.id, MatchScore.period == "FULL_TIME")
                        .first()
                    )
                    if not score:
                        score = MatchScore(
                            id=uuid4(),
                            match_id=match.id,
                            home_score=home_score,
                            away_score=away_score,
                            period="FULL_TIME",
                        )
                        db.add(score)
                    else:
                        score.home_score = home_score
                        score.away_score = away_score
                    db.commit()
            except Exception:
                pass
        except Exception as e:
            db.rollback()
            errors += 1

    return {"inserted": inserted, "updated": updated, "errors": errors}


def run_full_ingestion(
    db: Session,
    tennis_years: Optional[List[int]] = None,
    football_seasons: Optional[List[Tuple[str, str]]] = None,
    force_download: bool = False,
) -> Dict[str, Any]:
    """Run historical ingestion for configured years/leagues."""
    tennis_years = tennis_years or []
    football_seasons = football_seasons or []

    job = IngestionJob(
        id=uuid4(),
        job_type="full_sync",
        status="running",
        params={
            "tennis_years": tennis_years,
            "football_seasons": football_seasons,
        },
    )
    db.add(job)
    db.commit()
    db.refresh(job)

    summary = {"job_id": str(job.id), "tennis": {}, "football": {}}
    errors: List[str] = []

    try:
        for year in tennis_years:
            result = ingest_tennis_season(db, year, force_download=force_download)
            summary["tennis"][str(year)] = result

        for season, league in football_seasons:
            result = ingest_football_season(db, season, league, force_download=force_download)
            summary["football"][f"{season}_{league}"] = result

        job.status = "completed"
        job.result = summary
        db.commit()
    except Exception as e:
        job.status = "failed"
        job.result = {"error": str(e)}
        db.commit()
        raise

    return summary
